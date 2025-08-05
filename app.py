import os
from flask import Flask, request, jsonify,send_file
from werkzeug.utils import secure_filename
from markitdown import MarkItDown
import requests
import magic
from dotenv import load_dotenv
import json
import yaml
import pytesseract
from pdf2image import convert_from_path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from io import BytesIO
import datetime
from flask_cors import CORS


load_dotenv()


def load_config():
    with open("config.yml", "r") as config_file:
        return yaml.safe_load(config_file)


# Load configuration
config = load_config()
app = Flask(__name__)
ORIGIN = os.getenv("FRONT_URI")
CORS(app, origins=[ORIGIN])

app.config["MAX_CONTENT_LENGTH"] = int(config["general"]["max_file_size"])
app.config["UPLOAD_FOLDER"] = config["general"]["upload_folder"]

# Ensure upload folder exists
os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)

ALLOWED_EXTENSIONS = {"pdf", "docx"}
OPENROUTER_API_KEY = os.getenv("OPENROUTER_API_KEY")

def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def extract_markdown(file_path):
    md = MarkItDown()
    return str(md.convert(file_path).text_content)

def ocr_pdf(file_path):
    images = convert_from_path(file_path)
    text = ""
    for image in images:
        text += pytesseract.image_to_string(image)
    return text

def process_resume_with_ai(text):
    headers = {
        "Authorization": f"Bearer {OPENROUTER_API_KEY}",
        "Content-Type": "application/json",
        "HTTP-Referer": config["site"]["url"],
        "X-Title": config["site"]["name"],
    }

    prompt = f"""Please analyze this resume and extract information into a structured JSON object with the following fields exactly:
    - Name: 
    - Contact Number:
    - Email:
    - skills: list of key skills or technologies(comma seperated) 
    - Total Exp: gather data from experiences(if any) and calculate total (in years)
    - Current location: (if mentioned)
    - Cctc: Current ctc (if mentioned)
    - Ectc: Expected ctc (if mentioned)
    - Notice period: (if mentioned)

    Only include fields that are present in the resume. The output JSON must follow this exact structure with these field names (lists where appropriate). Provide valid JSON only, no extra text or explanation.

    Here is the resume text:

    {text}
    """

    response = requests.post(
        "https://openrouter.ai/api/v1/chat/completions",
        headers=headers,
        json={
            "models": config["models"][0:2],
            "messages": [{"role": "user", "content": prompt}],
            "response_format": {
                "type": "json_schema",
                "schema": config["resume_schema"],
            },
        },
    )

    if response.status_code != 200:
        print(response.json())
        raise Exception("Failed to process with AI")

    result = response.json()

    try:
        # Extract the JSON from the AI response
        content = result["choices"][0]["message"]["content"]
        # Find the JSON part in the response
        start_idx = content.find("{")
        end_idx = content.rfind("}") + 1
        if start_idx != -1 and end_idx != -1:
            json_str = content[start_idx:end_idx]
            return json.loads(json_str)
        return json.loads(content)
    except Exception as e:
        raise Exception(f"Failed to parse AI response: {str(e)}")

@app.route("/parse-multiple", methods=["POST"])
def parse_multiple_resumes():
    SECURE_API_KEY = os.getenv("SECURE_API_KEY", None)
    if SECURE_API_KEY is not None:
        api_key = request.headers.get('x-api-key')
        if api_key != SECURE_API_KEY or api_key is None:
            return jsonify({"error": "Unauthorized"}), 401

    # Check that 'files' key exists in files
    if "files" not in request.files:
        return jsonify({"error": "No files provided"}), 400

    files = request.files.getlist("files")
    if len(files) == 0:
        return jsonify({"error": "No files selected"}), 400

    results = []

    for file in files:
        if file.filename == "":
            continue
        if not allowed_file(file.filename):
            results.append({"filename": file.filename, "error": "Invalid file type"})
            continue

        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config["UPLOAD_FOLDER"], filename)
        file.save(file_path)

        try:
            file_type = magic.from_file(file_path, mime=True)

            if file_type == "application/pdf" or file_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
                text = extract_markdown(file_path)
                if file_type == "application/pdf" and (not text.strip() or len(text.strip()) < 10):
                    text = ocr_pdf(file_path)
            else:
                os.remove(file_path)
                results.append({"filename": filename, "error": "Unsupported file type"})
                continue

            result = process_resume_with_ai(text)

            os.remove(file_path)

            results.append({"filename": filename, "parsed_data": result})

        except Exception as e:
            if os.path.exists(file_path):
                os.remove(file_path)
            results.append({"filename": filename, "error": str(e)})

    
    columns = [
        "S no", "Date", "Name", "Contact number", "E mail id", "Skill", 
        "Current location", "Preferred location", "Total Exp", "Rel Exp", 
        "Cctc", "Ectc", "Notice period"
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Parsed Resumes"

    # Write header row
    for idx, col_name in enumerate(columns, start=1):
        ws.cell(row=1, column=idx, value=col_name)

    # Fill rows with parsed data
    for i, result in enumerate(results, start=1):
        row_num = i + 1  # Start from second row (first reserved for headers)
        ws.cell(row=row_num, column=1, value=i)  # S no
        ws.cell(row=row_num, column=2, value=datetime.date.today().strftime("%Y-%m-%d"))  # Date

        parsed = result.get("parsed_data", {})
        
        # Write other columns: use .get and place empty string if missing
        ws.cell(row=row_num, column=3, value=parsed.get("Name", ""))
        ws.cell(row=row_num, column=4, value=parsed.get("Contact Number", ""))
        ws.cell(row=row_num, column=5, value=parsed.get("Email", ""))
        ws.cell(row=row_num, column=6, value=parsed.get("skills", ""))
        ws.cell(row=row_num, column=7, value=parsed.get("Current location", ""))
        ws.cell(row=row_num, column=8, value=parsed.get("Preferred location", ""))
        ws.cell(row=row_num, column=9, value=parsed.get("Total Exp", ""))
        ws.cell(row=row_num, column=10, value=parsed.get("Rel Exp", ""))
        ws.cell(row=row_num, column=11, value=parsed.get("Cctc", ""))
        ws.cell(row=row_num, column=12, value=parsed.get("Ectc", ""))
        ws.cell(row=row_num, column=13, value=parsed.get("Notice period", ""))

    # Adjust column widths (optional)
    for col_idx, col_name in enumerate(columns, start=1):
        max_length = max(len(str(ws.cell(row=row, column=col_idx).value or "")) for row in range(1, ws.max_row + 1))
        adjusted_width = max_length + 2
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # Save to in-memory file
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Send file as response
    return send_file(
        output,
        as_attachment=True,
        download_name="parsed_resumes.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    

if __name__ == "__main__":
    app.run(debug=True)
